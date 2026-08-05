import csv
import io
from datetime import date as date_type

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.pdfencrypt import StandardEncryption
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlmodel.ext.asyncio.session import AsyncSession

from app.api.categories import list_categories
from app.api.cashback import get_cashback_summary
from app.api.dashboard import category_breakdown, monthly_overview, net_worth_dashboard, payment_method_breakdown
from app.api.goals import list_goals
from app.api.payment_methods import list_payment_methods
from app.api.recurring_bills import list_recurring_bills
from app.api.transactions import list_transactions
from app.models.user import User

_STYLES = getSampleStyleSheet()


async def build_transactions_csv(session: AsyncSession, user: User) -> bytes:
    transactions = await list_transactions(
        month=None,
        category_id=None,
        payment_method_id=None,
        amount_min=None,
        amount_max=None,
        search=None,
        transaction_type=None,
        limit=None,
        offset=0,
        user=user,
        session=session,
    )
    payment_methods = await list_payment_methods(user=user, session=session)
    categories = await list_categories(user=user, session=session)
    payment_method_names = {pm.id: pm.name for pm in payment_methods}
    category_names = {c.id: c.name for c in categories}

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        ["Date", "Merchant", "Description", "Type", "Source", "Payment Method", "Category", "Item", "Amount", "Notes"]
    )
    for transaction in transactions:
        payment_method_name = payment_method_names.get(transaction.payment_method_id, "")
        for item in transaction.line_items:
            writer.writerow(
                [
                    transaction.date.isoformat(),
                    transaction.merchant,
                    transaction.description or "",
                    transaction.transaction_type,
                    transaction.source,
                    payment_method_name,
                    category_names.get(item.category_id, ""),
                    item.item_name,
                    str(item.amount),
                    transaction.notes or "",
                ]
            )

    # utf-8-sig prepends the BOM as bytes (not a literal character before
    # encoding) so Excel detects the encoding correctly while strict CSV
    # readers that decode with utf-8-sig still see a clean first cell.
    return buffer.getvalue().encode("utf-8-sig")


async def build_monthly_report_xlsx(session: AsyncSession, user: User, month: int, year: int) -> bytes:
    overview = await monthly_overview(month=month, year=year, user=user, session=session)
    categories = await category_breakdown(month=month, year=year, user=user, session=session)
    payment_methods = await payment_method_breakdown(month=month, year=year, user=user, session=session)
    cashback = await get_cashback_summary(year=year, month=month, user=user, session=session)
    recurring_bills = await list_recurring_bills(user=user, session=session)
    goals = await list_goals(user=user, session=session)
    net_worth = await net_worth_dashboard(user=user, session=session)
    transactions = await list_transactions(
        month=f"{year:04d}-{month:02d}",
        category_id=None,
        payment_method_id=None,
        amount_min=None,
        amount_max=None,
        search=None,
        transaction_type=None,
        limit=None,
        offset=0,
        user=user,
        session=session,
    )
    payment_method_names = {pm.id: pm.name for pm in await list_payment_methods(user=user, session=session)}
    category_names = {c.id: c.name for c in await list_categories(user=user, session=session)}

    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Month", f"{year}-{month:02d}"])
    summary.append(["Total Income", str(overview.total_income)])
    summary.append(["Total Expenses", str(overview.total_expenses)])
    summary.append(["Net Cash Flow", str(overview.net_cash_flow)])
    summary.append(["Top Category", overview.top_category.name if overview.top_category else ""])
    summary.append(["Top Payment Method", overview.top_payment_method.name if overview.top_payment_method else ""])
    summary.append(
        [
            "Change vs Previous Month",
            str(overview.comparison_vs_previous_month.change_amount),
        ]
    )

    txn_sheet = workbook.create_sheet("Transactions")
    txn_sheet.append(["Date", "Merchant", "Type", "Payment Method", "Category", "Item", "Amount", "Notes"])
    for transaction in transactions:
        payment_method_name = payment_method_names.get(transaction.payment_method_id, "")
        for item in transaction.line_items:
            txn_sheet.append(
                [
                    transaction.date.isoformat(),
                    transaction.merchant,
                    transaction.transaction_type,
                    payment_method_name,
                    category_names.get(item.category_id, ""),
                    item.item_name,
                    str(item.amount),
                    transaction.notes or "",
                ]
            )

    category_sheet = workbook.create_sheet("Category Breakdown")
    category_sheet.append(["Category", "Amount", "% of Total", "Budget", "% of Budget Used", "Over Budget"])
    for item in categories:
        category_sheet.append(
            [
                item.name,
                str(item.amount),
                str(item.percentage_of_total),
                str(item.budget_amount) if item.budget_amount is not None else "",
                str(item.budget_percentage_used) if item.budget_percentage_used is not None else "",
                "Yes" if item.is_over_budget else "No",
            ]
        )

    budget_sheet = workbook.create_sheet("Budget vs Actual")
    budget_sheet.append(["Category", "Budget", "Actual", "% Used", "Over Budget"])
    for item in categories:
        if item.budget_amount is None:
            continue
        budget_sheet.append(
            [
                item.name,
                str(item.budget_amount),
                str(item.amount),
                str(item.budget_percentage_used) if item.budget_percentage_used is not None else "",
                "Yes" if item.is_over_budget else "No",
            ]
        )

    pm_sheet = workbook.create_sheet("Payment Methods")
    pm_sheet.append(["Payment Method", "Type", "Amount", "Transaction Count", "Average Transaction"])
    for item in payment_methods:
        pm_sheet.append([item.name, item.type, str(item.amount), item.transaction_count, str(item.average_transaction)])

    cashback_sheet = workbook.create_sheet("Cashback")
    cashback_sheet.append(["Total Estimated", str(cashback.total_estimated)])
    cashback_sheet.append(["Total Redeemed", str(cashback.total_redeemed)])
    cashback_sheet.append(["Total Unredeemed", str(cashback.total_unredeemed)])
    cashback_sheet.append([])
    cashback_sheet.append(["By Card", "Estimated", "Redeemed"])
    for card in cashback.by_card:
        cashback_sheet.append([card.name, str(card.estimated), str(card.redeemed)])
    cashback_sheet.append([])
    cashback_sheet.append(["By Category", "Estimated", "Redeemed"])
    for category in cashback.by_category:
        cashback_sheet.append([category.name, str(category.estimated), str(category.redeemed)])

    bills_sheet = workbook.create_sheet("Recurring Bills")
    bills_sheet.append(["Name", "Amount", "Frequency", "Due Date", "Active", "Current Period Status"])
    for bill in recurring_bills:
        bills_sheet.append(
            [
                bill.name,
                str(bill.amount),
                bill.frequency,
                bill.due_date.isoformat(),
                "Yes" if bill.is_active else "No",
                bill.current_period.status if bill.current_period else "",
            ]
        )

    net_worth_sheet = workbook.create_sheet("Net Worth")
    net_worth_sheet.append(["Current Net Worth", str(net_worth.current_net_worth) if net_worth.current_net_worth is not None else ""])
    net_worth_sheet.append(["Total Assets", str(net_worth.total_assets) if net_worth.total_assets is not None else ""])
    net_worth_sheet.append(
        ["Total Liabilities", str(net_worth.total_liabilities) if net_worth.total_liabilities is not None else ""]
    )
    net_worth_sheet.append([])
    net_worth_sheet.append(["Account", "Type", "Balance"])
    for balance in net_worth.breakdown:
        net_worth_sheet.append([balance.account_name, balance.account_type, str(balance.balance)])

    goals_sheet = workbook.create_sheet("Goals")
    goals_sheet.append(["Name", "Target Amount", "Current Amount", "Target Date", "Shared", "Active"])
    for goal in goals:
        goals_sheet.append(
            [
                goal.name,
                str(goal.target_amount),
                str(goal.current_amount),
                goal.target_date.isoformat() if goal.target_date else "",
                "Yes" if goal.is_shared else "No",
                "Yes" if goal.is_active else "No",
            ]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _table(rows: list[list[str]], col_widths: list[float] | None = None) -> Table:
    table = Table(rows, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dddddd")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


async def build_monthly_report_pdf(
    session: AsyncSession, user: User, month: int, year: int, password: str | None
) -> bytes:
    overview = await monthly_overview(month=month, year=year, user=user, session=session)
    categories = await category_breakdown(month=month, year=year, user=user, session=session)
    payment_methods = await payment_method_breakdown(month=month, year=year, user=user, session=session)
    cashback = await get_cashback_summary(year=year, month=month, user=user, session=session)
    recurring_bills = await list_recurring_bills(user=user, session=session)
    goals = await list_goals(user=user, session=session)
    net_worth = await net_worth_dashboard(user=user, session=session)

    story = [
        Paragraph(f"BillWise Monthly Report — {date_type(year, month, 1).strftime('%B %Y')}", _STYLES["Title"]),
        Spacer(1, 0.2 * inch),
        Paragraph(
            f"Income: ${overview.total_income} &nbsp;&nbsp; Expenses: ${overview.total_expenses} "
            f"&nbsp;&nbsp; Net cash flow: ${overview.net_cash_flow}",
            _STYLES["Normal"],
        ),
        Spacer(1, 0.2 * inch),
        Paragraph("Category Breakdown", _STYLES["Heading2"]),
        _table(
            [["Category", "Amount", "% of Total", "Over Budget"]]
            + [[c.name, str(c.amount), f"{c.percentage_of_total}%", "Yes" if c.is_over_budget else "No"] for c in categories]
        ),
        Spacer(1, 0.2 * inch),
        Paragraph("Payment Method Breakdown", _STYLES["Heading2"]),
        _table(
            [["Payment Method", "Amount", "Transactions"]]
            + [[p.name, str(p.amount), str(p.transaction_count)] for p in payment_methods]
        ),
        Spacer(1, 0.2 * inch),
        Paragraph("Cashback Summary", _STYLES["Heading2"]),
        Paragraph(
            f"Estimated: ${cashback.total_estimated} &nbsp;&nbsp; Redeemed: ${cashback.total_redeemed} "
            f"&nbsp;&nbsp; Unredeemed: ${cashback.total_unredeemed}",
            _STYLES["Normal"],
        ),
        Spacer(1, 0.2 * inch),
        Paragraph("Recurring Bills", _STYLES["Heading2"]),
        _table(
            [["Name", "Amount", "Frequency", "Due Date"]]
            + [[b.name, str(b.amount), b.frequency, b.due_date.isoformat()] for b in recurring_bills]
        ),
        Spacer(1, 0.2 * inch),
        Paragraph("Goals Progress", _STYLES["Heading2"]),
        _table(
            [["Goal", "Target", "Current"]]
            + [[g.name, str(g.target_amount), str(g.current_amount)] for g in goals]
        ),
        Spacer(1, 0.2 * inch),
        Paragraph("Net Worth Snapshot", _STYLES["Heading2"]),
        Paragraph(
            f"Net worth: ${net_worth.current_net_worth if net_worth.current_net_worth is not None else 'N/A'} "
            f"&nbsp;&nbsp; Assets: ${net_worth.total_assets if net_worth.total_assets is not None else 'N/A'} "
            f"&nbsp;&nbsp; Liabilities: ${net_worth.total_liabilities if net_worth.total_liabilities is not None else 'N/A'}",
            _STYLES["Normal"],
        ),
    ]

    buffer = io.BytesIO()
    encrypt = StandardEncryption(userPassword=password, ownerPassword=password) if password else None
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, encrypt=encrypt)
    doc.build(story)
    return buffer.getvalue()
